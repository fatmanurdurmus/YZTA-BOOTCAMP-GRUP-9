"""Safe, in-memory text extraction for supported uploaded documents."""

from io import BytesIO


SUPPORTED_ACTIVITY_TYPES = {"application/pdf", "application/vnd.openxmlformats-officedocument.wordprocessingml.document", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
SUPPORTED_LAW_TYPES = SUPPORTED_ACTIVITY_TYPES | {"text/plain"}


class UnsupportedDocumentType(ValueError):
    pass


def extract_document_text(content: bytes, content_type: str, filename: str) -> list[tuple[int, str]]:
    """Return page/sheet numbered text without persisting the uploaded file."""
    normalized_type = content_type or _type_from_filename(filename)
    if normalized_type == "application/pdf":
        from pypdf import PdfReader

        return [(number, page.extract_text() or "") for number, page in enumerate(PdfReader(BytesIO(content)).pages, 1)]
    if normalized_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
        from docx import Document

        document = Document(BytesIO(content))
        return [(1, "\n".join(paragraph.text for paragraph in document.paragraphs))]
    if normalized_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        return [
            (number, "\n".join(" | ".join(str(cell) for cell in row if cell is not None) for row in sheet.iter_rows(values_only=True)))
            for number, sheet in enumerate(workbook.worksheets, 1)
        ]
    if normalized_type == "text/plain":
        return [(1, content.decode("utf-8"))]
    raise UnsupportedDocumentType(f"Unsupported document type: {content_type or filename}")


def _type_from_filename(filename: str) -> str:
    suffix = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    return {"pdf": "application/pdf", "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document", "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "txt": "text/plain"}.get(suffix, "")
